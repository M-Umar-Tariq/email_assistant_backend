"""
Extract readable text from email attachments (PDF, DOCX, XLSX, TXT, CSV, etc.).
"""

import io
import csv
import traceback


SUPPORTED_MIMETYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/plain",
    "text/csv",
    "text/html",
    "application/rtf",
}

MAX_TEXT_PER_ATTACHMENT = 8000


def is_supported(content_type: str) -> bool:
    return content_type.lower().split(";")[0].strip() in SUPPORTED_MIMETYPES


def _extract_pdf(data: bytes) -> str:
    import fitz  # PyMuPDF
    try:
        with fitz.open(stream=data, filetype="pdf") as doc:
            if doc.is_encrypted:
                return ""
            text_parts = []
            for page in doc:
                text_parts.append(page.get_text())
        return "\n".join(text_parts)
    except (ValueError, RuntimeError):
        # Encrypted, closed, corrupted, or otherwise unreadable PDF
        return ""


def _extract_docx(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"[Sheet: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append(" | ".join(cells))
    wb.close()
    return "\n".join(lines)


def _extract_csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines = []
    for row in reader:
        lines.append(" | ".join(row))
    return "\n".join(lines)


def _extract_plain_text(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


_EXTRACTOR_MAP = {
    "application/pdf": _extract_pdf,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": _extract_docx,
    "application/msword": _extract_docx,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": _extract_xlsx,
    "application/vnd.ms-excel": _extract_xlsx,
    "text/plain": _extract_plain_text,
    "text/csv": _extract_csv,
    "text/html": _extract_plain_text,
    "application/rtf": _extract_plain_text,
}


def extract_text(content_type: str, data: bytes) -> str:
    """Extract text from attachment binary data. Returns empty string on failure."""
    ct = content_type.lower().split(";")[0].strip()
    extractor = _EXTRACTOR_MAP.get(ct)
    if not extractor:
        return ""
    try:
        text = extractor(data)
        return text[:MAX_TEXT_PER_ATTACHMENT] if len(text) > MAX_TEXT_PER_ATTACHMENT else text
    except Exception:
        traceback.print_exc()
        return ""


def extract_attachments_from_message(msg, include_binary: bool = True) -> list[dict]:
    """
    Walk an email.message.Message and extract attachments.
    Returns list of dicts with filename, content_type, text, size, and optionally raw binary data.
    """
    import base64 as b64mod

    attachments = []
    if not msg.is_multipart():
        return attachments

    for part in msg.walk():
        disposition = part.get_content_disposition()
        if disposition not in ("attachment", "inline"):
            continue

        ct = part.get_content_type() or ""
        filename = part.get_filename() or "unnamed"

        if ct.startswith("image/") and disposition != "attachment":
            continue

        payload = part.get_payload(decode=True)
        if not payload:
            continue

        size = len(payload)
        text = ""
        if is_supported(ct):
            text = extract_text(ct, payload)

        entry = {
            "filename": filename,
            "content_type": ct,
            "text": text,
            "size": size,
        }
        if include_binary:
            entry["data_b64"] = b64mod.b64encode(payload).decode("ascii")

        attachments.append(entry)

    return attachments
